from rest_framework import viewsets, views, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import date
from decimal import Decimal
import uuid

from .models import (
    Usuario, Musico, Evento, Asistencia, Descuento, Pago,
    RendimientoMusico, ConfiguracionSistema, Adelanto, PlanillaLiquidacion,
    ContratoMusico, DetalleMontoDiario, JefeSeccion, Deuda, AbonoDeuda
)
from .models import Modulo, RolModulo, UsuarioModulo
from .serializers import (
    UsuarioSerializer, UsuarioCreateSerializer, ModuloSerializer, RolModuloSerializer,
    MusicoSerializer, EventoSerializer,
    AsistenciaSerializer, DescuentoSerializer, PagoSerializer,
    ConfiguracionSistemaSerializer, AdelantoSerializer,
    PlanillaLiquidacionSerializer, PlanillaLiquidacionDetalleSerializer,
)
from services.rendimiento_calculator import RendimientoCalculator

class UserMeView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        serializer = UsuarioSerializer(request.user)
        data = serializer.data
        if hasattr(request.user, 'perfil_musico'):
            data['musico_id'] = request.user.perfil_musico.id
        return Response(data)

class DashboardView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        
        if user.rol in ['DIRECTOR', 'SUBDIRECTOR']:
            total_musicos = Musico.objects.filter(activo=True).count()
            total_eventos = Evento.objects.count()
            eventos_mes = Evento.objects.filter(
                fecha_hora_cita__month=date.today().month,
                fecha_hora_cita__year=date.today().year
            ).count()
            
            calculator = RendimientoCalculator()
            top_musicos = calculator.obtener_top_musicos(limite=5)
            eventos_recientes = Evento.objects.order_by('-fecha_hora_cita')[:5]
            
            return Response({
                'rol': user.rol,
                'total_musicos': total_musicos,
                'total_eventos': total_eventos,
                'eventos_mes': eventos_mes,
                'top_musicos': top_musicos,
                'eventos_recientes': EventoSerializer(eventos_recientes, many=True).data,
            })
        else:
            try:
                musico = request.user.perfil_musico
                proximos_eventos = Evento.objects.filter(
                    fecha_hora_cita__gte=timezone.now()
                ).order_by('fecha_hora_cita')[:5]
                asistencias = Asistencia.objects.filter(musico=musico).order_by('-evento__fecha_hora_cita')[:5]
                pagos = Pago.objects.filter(musico=musico).order_by('-fecha_liquidacion')[:5]
                
                return Response({
                    'rol': user.rol,
                    'proximos_eventos': EventoSerializer(proximos_eventos, many=True).data,
                    'asistencias': AsistenciaSerializer(asistencias, many=True).data,
                    'pagos': PagoSerializer(pagos, many=True).data,
                })
            except Musico.DoesNotExist:
                return Response({'error': 'No tienes perfil de músico'}, status=status.HTTP_400_BAD_REQUEST)

class MusicoViewSet(viewsets.ModelViewSet):
    queryset = Musico.objects.filter(activo=True).order_by('orden', 'apellidos', 'nombres')
    serializer_class = MusicoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        instrumento = self.request.query_params.get('instrumento', None)
        
        if search:
            queryset = queryset.filter(
                Q(nombres__icontains=search) | Q(apellidos__icontains=search) | Q(documento_identidad__icontains=search)
            )
        if instrumento:
            queryset = queryset.filter(instrumento=instrumento)
            
        return queryset

    def perform_create(self, serializer):
        data = self.request.data
        nombres = data.get('nombres', '')
        apellidos = data.get('apellidos', '')
        ci = data.get('documento_identidad', '')
        
        base_username = nombres.split()[0].lower() if nombres else 'musico'
        username = f"{base_username}_{uuid.uuid4().hex[:6]}"
        password = ci if ci else '123456'
        
        user = Usuario.objects.create_user(
            username=username,
            password=password,
            first_name=nombres,
            last_name=apellidos,
            rol='MUSICO'
        )
        serializer.save(usuario=user)

    @action(detail=False, methods=['post'])
    def update_order(self, request):
        order_data = request.data.get('order', [])
        for item in order_data:
            Musico.objects.filter(id=item['id']).update(orden=item['order'])
        return Response({'success': True})

    @action(detail=False, methods=['get'])
    def generar_plantilla_excel(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Descuentos"
        
        # ---------------------------------------------------------
        # CONFIGURACIÓN DE PÁGINA PARA PDF
        # ---------------------------------------------------------
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.3937
        ws.page_margins.right = 0.3937
        ws.page_margins.top = 0.3937
        ws.page_margins.bottom = 0.3937
        ws.page_margins.header = 0.3937
        ws.page_margins.footer = 0.3937
        ws.print_options.horizontalCentered = True
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        main_title_font = Font(name='Arial', size=16, bold=True, color="000000")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        section_font = Font(name='Arial', size=10, bold=True, color="000000")
        data_font = Font(name='Arial', size=9, color="000000")
        total_font = Font(name='Arial', size=9, bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        
        ws.merge_cells('A1:G2')
        ws['A1'] = "Descuento Pdf"
        ws['A1'].font = main_title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['N°', 'NOMBRES Y APELLIDOS', 'ATRASO', 'FALTA', 'UNIFORME', 'BEBIDAS', 'TOTAL']
        ws.row_dimensions[4].height = 80 
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            if col_num > 2:
                cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 6
        
        musicos = Musico.objects.filter(activo=True).order_by('orden', 'apellidos')
        secciones = ['TROMPETA', 'CLARINETE', 'SAXOFON', 'BARITONO', 'TROMBON', 'TUBA', 'BOMBO', 'TAMBOR', 'PLATILLOS', 'PERCUSION', 'OTRO']
        current_row = 5
        for seccion in secciones:
            musicos_seccion = [m for m in musicos if m.instrumento == seccion]
            if not musicos_seccion:
                continue
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            cell = ws.cell(row=current_row, column=1, value=seccion)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
            for idx, musico in enumerate(musicos_seccion, 1):
                c1 = ws.cell(row=current_row, column=1, value=idx)
                c1.border = thin_border
                c1.font = data_font
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c2 = ws.cell(row=current_row, column=2, value=musico.nombre_completo)
                c2.border = thin_border
                c2.font = data_font
                c2.alignment = Alignment(horizontal="left", vertical="center")
                for col in range(3, 8):
                    cell = ws.cell(row=current_row, column=col, value="")
                    cell.border = thin_border
                current_row += 1
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_descuentos.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def reporte(self, request):
        formato = request.query_params.get('tipo_reporte', 'excel')
        columns_str = request.query_params.get('columns', '')
        columns = columns_str.split(',') if columns_str else ['nombres_apellidos', 'instrumento']

        musicos = self.get_queryset()

        from django.conf import settings
        import os
        logo_path = os.path.join(settings.BASE_DIR, 'media', 'imagenes', 'imagen_logo.jpg')

        if formato == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.drawing.image import Image as ExcelImage
            except ImportError:
                return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Músicos"

            titulo_personalizado = request.query_params.get('titulo', '')

            # --- DISEÑO DEL ENCABEZADO EXCEL ---
            ws.merge_cells('A1:H2')
            ws['A1'] = "BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA\nEco De Los Andes"
            ws['A1'].font = Font(bold=True, size=14, color="000000")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if titulo_personalizado:
                ws.merge_cells('A3:H3')
                ws['A3'] = titulo_personalizado.upper()
                ws['A3'].font = Font(bold=True, size=12, color="000000")
                ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('A4:H4')
            ws['A4'] = '"Por que la meji nunca pierde papá "'
            ws['A4'].font = Font(italic=True, size=11, color="000000")
            ws['A4'].alignment = Alignment(horizontal="center", vertical="center")

            # Espacio
            start_row = 7

            raw_headers = [col.replace('_', ' ').upper() for col in columns]
            headers = ['TALLAS (C/Ch/Z)' if h == 'TALLAS' else h for h in raw_headers]

            # Estilos de cabecera tabla
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for col_num, header_text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            current_row = start_row + 1
            for musico in musicos:
                row = []
                for col in columns:
                    if col == 'ci':
                        row.append(musico.documento_identidad or '')
                    elif col == 'nombres_apellidos':
                        row.append(f"{musico.nombres} {musico.apellidos}")
                    elif col == 'celular':
                        row.append(musico.telefono or '')
                    elif col == 'instrumento':
                        row.append(musico.instrumento or '')
                    elif col == 'tallas':
                        row.append(f"{musico.talla_camisa or '-'}/{musico.talla_chamarra or '-'}/{musico.numero_calzado or '-'}")
                    elif col == 'estado':
                        row.append('Activo' if musico.activo else 'Inactivo')
                    elif col == 'direccion':
                        row.append(musico.direccion or '')
                    elif col == 'fecha_nacimiento':
                        row.append(str(musico.fecha_nacimiento) if musico.fecha_nacimiento else '')
                    elif col == 'nivel':
                        row.append(musico.nivel or '')
                    else:
                        row.append('')

                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=cell_value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="left")
                current_row += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Musicos.xlsx'
            wb.save(response)
            return response

        elif formato == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, portrait
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            except ImportError:
                return Response({'error': 'reportlab no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            import io
            from django.http import HttpResponse

            buffer = io.BytesIO()
            # Hoja Vertical (Portrait)
            doc = SimpleDocTemplate(buffer, pagesize=portrait(letter),
                                    rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
            elements = []

            styles = getSampleStyleSheet()

            titulo_personalizado = request.query_params.get('titulo', '')

            # --- DISEÑO DEL ENCABEZADO PDF ---
            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
                fontSize=14, textColor=colors.black, alignment=TA_CENTER,
                spaceAfter=5
            )
            custom_title_style = ParagraphStyle(
                'UserTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
                fontSize=12, textColor=colors.black, alignment=TA_CENTER,
                spaceAfter=10
            )
            phrase_style = ParagraphStyle(
                'PhraseStyle', parent=styles['Normal'], fontName='Helvetica-Oblique',
                fontSize=10, textColor=colors.black, alignment=TA_CENTER,
                spaceAfter=20
            )

            elements.append(Paragraph("BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA<br/>Eco De Los Andes", title_style))
            if titulo_personalizado:
                elements.append(Paragraph(titulo_personalizado.upper(), custom_title_style))

            elements.append(Paragraph('"Por que la meji nunca pierde papá "', phrase_style))
            elements.append(Spacer(1, 10))

            # Tabla de datos
            raw_headers = [col.replace('_', ' ').upper() for col in columns]
            headers = ['TALLAS (C/Ch/Z)' if h == 'TALLAS' else h for h in raw_headers]
            data = [headers]

            # Estilo más pequeño para que quepa en vertical
            small_style = ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=8)

            for musico in musicos:
                row = []
                for col in columns:
                    if col == 'ci':
                        row.append(musico.documento_identidad or '')
                    elif col == 'nombres_apellidos':
                        row.append(f"{musico.nombres} {musico.apellidos}")
                    elif col == 'celular':
                        row.append(musico.telefono or '')
                    elif col == 'instrumento':
                        row.append(musico.instrumento or '')
                    elif col == 'tallas':
                        row.append(f"{musico.talla_camisa or '-'}/{musico.talla_chamarra or '-'}/{musico.numero_calzado or '-'}")
                    elif col == 'estado':
                        row.append('Activo' if musico.activo else 'Inactivo')
                    elif col == 'direccion':
                        row.append(musico.direccion or '')
                    elif col == 'fecha_nacimiento':
                        row.append(str(musico.fecha_nacimiento) if musico.fecha_nacimiento else '')
                    elif col == 'nivel':
                        row.append(musico.nivel or '')
                    else:
                        row.append('')

                data.append([Paragraph(str(item), small_style) if len(str(item)) > 15 else str(item) for item in row])

            # Reduce font size to fit portrait
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E5E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(t)

            # Función para dibujar el pie de página
            def footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 8)
                canvas.drawRightString(doc.pagesize[0] - 30, 20, f"Página {doc.page}")
                canvas.restoreState()

            doc.build(elements, onFirstPage=footer, onLaterPages=footer)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Musicos.pdf'
            return response

        return Response({'error': 'Formato no soportado'}, status=status.HTTP_400_BAD_REQUEST)


class PresidentePermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.is_superuser or request.user.rol == 'PRESIDENTE'


class ModuloViewSet(viewsets.ModelViewSet):
    queryset = Modulo.objects.all().order_by('clave')
    serializer_class = ModuloSerializer
    permission_classes = [permissions.IsAuthenticated, PresidentePermission]


class RolModuloViewSet(viewsets.ModelViewSet):
    queryset = RolModulo.objects.select_related('modulo').all().order_by('rol', 'modulo__clave')
    serializer_class = RolModuloSerializer
    permission_classes = [permissions.IsAuthenticated, PresidentePermission]


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all().order_by('rol', 'last_name', 'first_name')
    permission_classes = [permissions.IsAuthenticated, PresidentePermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return UsuarioCreateSerializer
        return UsuarioSerializer

    @action(detail=False, methods=['post'])
    def update_order(self, request):
        order_data = request.data.get('order', [])
        for item in order_data:
            Musico.objects.filter(id=item['id']).update(orden=item['order'])
        return Response({'success': True})

    @action(detail=False, methods=['get'])
    def generar_plantilla_excel(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Descuentos"
        
        # ---------------------------------------------------------
        # CONFIGURACIÓN DE PÁGINA PARA PDF
        # ---------------------------------------------------------
        # Orientación Horizontal (Landscape) para tener el doble de espacio horizontal
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        
        # Márgenes exactos de 1 cm (0.3937 pulgadas) en todos los lados
        ws.page_margins.left = 0.3937
        ws.page_margins.right = 0.3937
        ws.page_margins.top = 0.3937
        ws.page_margins.bottom = 0.3937
        ws.page_margins.header = 0.3937
        ws.page_margins.footer = 0.3937
        
        # Centrar la tabla horizontalmente en la hoja
        ws.print_options.horizontalCentered = True
        
        # Auto-escalado: Encoge todo lo necesario para que el ancho entre en 1 sola página
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # ---------------------------------------------------------
        # ESTILOS PROFESIONALES
        # ---------------------------------------------------------
        # Fuentes principales
        main_title_font = Font(name='Arial', size=16, bold=True, color="000000")
        
        # Estilos de Cabecera (Fondo Azul Oscuro Institucional, Letra Blanca)
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        
        # Estilos de Sección (Fondo Gris Claro, Letra Negra)
        section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        section_font = Font(name='Arial', size=10, bold=True, color="000000")
        
        # Estilos de Datos
        data_font = Font(name='Arial', size=9, color="000000")
        total_font = Font(name='Arial', size=9, bold=True, color="000000")
        
        # Bordes Negros Delgados bien definidos (Obligatorio para que no falle el lector PDF)
        thin_border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        
        # ---------------------------------------------------------
        # TÍTULO PRINCIPAL
        # ---------------------------------------------------------
        ws.merge_cells('A1:G2')
        ws['A1'] = "Descuento Pdf"
        ws['A1'].font = main_title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # ---------------------------------------------------------
        # CABECERAS Y ROTACIÓN
        # ---------------------------------------------------------
        headers = ['N°', 'NOMBRES Y APELLIDOS', 'ATRASO', 'FALTA', 'UNIFORME', 'BEBIDAS', 'TOTAL']
        
        # Aumentamos el alto de la fila para que el texto hacia arriba tenga espacio
        ws.row_dimensions[4].height = 80 
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            
            # Si son columnas de descuento o total (de la 3 en adelante), rotamos 90 grados y permitimos salto de línea
            if col_num > 2:
                cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # ---------------------------------------------------------
        # ANCHOS DE COLUMNA
        # ---------------------------------------------------------
        ws.column_dimensions['A'].width = 4   # N°
        ws.column_dimensions['B'].width = 30  # Nombres (Espacio holgado para nombres largos)
        
        # Las columnas de descuentos y total ahora son ajustadas, ancho 6 permite dos líneas de texto rotado
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 6
            
        # ---------------------------------------------------------
        # DATOS DE LOS MÚSICOS
        # ---------------------------------------------------------
        musicos = Musico.objects.filter(activo=True).order_by('orden', 'apellidos')
        secciones = ['TROMPETA', 'CLARINETE', 'SAXOFON', 'BARITONO', 'TROMBON', 'TUBA', 'BOMBO', 'TAMBOR', 'PLATILLOS', 'PERCUSION', 'OTRO']
        
        current_row = 5
        for seccion in secciones:
            musicos_seccion = [m for m in musicos if m.instrumento == seccion]
            if not musicos_seccion: continue
            
            # Fila separadora de sección
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            cell = ws.cell(row=current_row, column=1, value=seccion)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            # Borde para la celda combinada
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
                
            current_row += 1
            
            for idx, musico in enumerate(musicos_seccion, 1):
                # N°
                c1 = ws.cell(row=current_row, column=1, value=idx)
                c1.border = thin_border
                c1.font = data_font
                c1.alignment = Alignment(horizontal="center", vertical="center")
                
                # Nombres
                c2 = ws.cell(row=current_row, column=2, value=musico.nombre_completo)
                c2.border = thin_border
                c2.font = data_font
                c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                # Espacios vacíos para ingresar números
                for col in range(3, 7):
                    cv = ws.cell(row=current_row, column=col, value='')
                    cv.border = thin_border
                    cv.font = data_font
                    cv.alignment = Alignment(horizontal="center", vertical="center")
                
                # TOTAL (Fórmula en negrita)
                ctot = ws.cell(row=current_row, column=7, value=f"=SUM(C{current_row}:F{current_row})")
                ctot.font = total_font
                ctot.border = thin_border
                ctot.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
                
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Multas_Adelantos.xlsx'
        wb.save(response)
        return response


class EventoViewSet(viewsets.ModelViewSet):
    queryset = Evento.objects.all().order_by('-fecha_hora_cita')
    serializer_class = EventoSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=['post'])
    def registrar_asistencia(self, request, pk=None):
        evento = self.get_object()
        asistencias_data = request.data.get('asistencias', [])
        
        config = ConfiguracionSistema.objects.first()
        hora_limite = config.hora_limite_tardanza if config else evento.fecha_hora_cita.time()
        
        for item in asistencias_data:
            musico_id = item.get('musico_id')
            estado = item.get('estado', 'AUSENTE')
            hora_llegada = item.get('hora_llegada')
                
            Asistencia.objects.update_or_create(
                musico_id=musico_id,
                evento=evento,
                defaults={
                    'estado': estado,
                    'hora_llegada': hora_llegada if hora_llegada else None
                }
            )
        return Response({'success': True})

    @action(detail=True, methods=['post'])
    def generar_pagos(self, request, pk=None):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        evento = self.get_object()
        config = ConfiguracionSistema.objects.first()
        monto_base = config.monto_por_evento if config else Decimal('100.00')
        
        asistencias = Asistencia.objects.filter(
            evento=evento, estado__in=['PRESENTE', 'TARDANZA', 'JUSTIFICADO']
        )
        
        generados = 0
        for asistencia in asistencias:
            if not Pago.objects.filter(musico=asistencia.musico, evento=evento).exists():
                total_descuentos = Descuento.objects.filter(
                    musico=asistencia.musico
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
                
                Pago.objects.create(
                    musico=asistencia.musico,
                    evento=evento,
                    salario_base=monto_base,
                    descuentos_totales=total_descuentos,
                    adelantos_totales=Decimal('0.00'),
                    neto_pagar=monto_base - total_descuentos,
                    fecha_liquidacion=timezone.now(),
                    estado='PENDIENTE'
                )
                generados += 1
        return Response({'success': True, 'generados': generados})

    @action(detail=True, methods=['get'])
    def generar_mensaje(self, request, pk=None):
        evento = self.get_object()
        
        grupos = [
            ('TROMPETAS', ['TROMPETA']),
            ('SAXOS', ['SAXOFON']),
            ('CLARINETES', ['CLARINETE']),
            ('BARÍTONOS', ['BARITONO']),
            ('TROMBONES', ['TROMBON']),
            ('TUBAS', ['TUBA']),
            ('SECCIÓN PERCUSIÓN (BOMBOS, TAMBORES Y PLATILLOS)', ['BOMBO', 'TAMBOR', 'PLATILLOS', 'PERCUSION']),
        ]
        
        fecha_hora_local = timezone.localtime(evento.fecha_hora_cita)
        fecha_str = fecha_hora_local.strftime('%d/%m/%Y')
        hora_str = fecha_hora_local.strftime('%H:%M')
        
        lines = []
        lines.append("*RELACIÓN NOMINAL*")
        lines.append(f"*Evento:* {evento.titulo}")
        lines.append("")
        lines.append(f"*Fecha y Hora:* {fecha_str} - {hora_str} hrs")
        if evento.lugar_concentracion:
            lines.append(f"*Concentración:* {evento.lugar_concentracion}")
        
        uniforme_display = dict(Evento.UNIFORMES).get(evento.uniforme, evento.uniforme)
        if evento.uniforme == 'OTRO' and evento.uniforme_personalizado:
            uniforme_display = evento.uniforme_personalizado
            
        lines.append(f"*Uniforme:* {uniforme_display}")
        if evento.detalles_uniforme:
            lines.append(f"   _{evento.detalles_uniforme}_")
        lines.append("")
        
        convocados = evento.convocados.all()
        for titulo, instrumentos in grupos:
            musicos_seccion = [m for m in convocados if m.instrumento in instrumentos]
            if musicos_seccion:
                lines.append(f"*{titulo}*")
                for idx, musico in enumerate(musicos_seccion, 1):
                    lines.append(f"{idx}. {musico.nombres} {musico.apellidos}")
                lines.append("")
                
        lines.append("*Nota:* Se ruega puntualidad a la hora de concentración.")
        
        return Response({'mensaje': '\n'.join(lines)})

class AsistenciaViewSet(viewsets.ModelViewSet):
    queryset = Asistencia.objects.all()
    serializer_class = AsistenciaSerializer
    permission_classes = [permissions.IsAuthenticated]

class DescuentoViewSet(viewsets.ModelViewSet):
    queryset = Descuento.objects.all().order_by('-fecha_falta')
    serializer_class = DescuentoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrar descuentos según rol del usuario"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # DIRECTOR, SUBDIRECTOR, PRESIDENTE: ven todo
        if user.rol in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] or user.is_superuser:
            return queryset
        
        # JEFE DE SECCIÓN: solo descuentos de su sección
        if user.rol == 'JEFE_SECCION':
            try:
                jefe = JefeSeccion.objects.get(usuario=user)
                # Filtrar por sección del jefe
                return queryset.filter(musico__instrumento=jefe.seccion)
            except JefeSeccion.DoesNotExist:
                return queryset.none()
        
        # Otros roles: no ven descuentos
        return queryset.none()



    @action(detail=False, methods=['post'])
    def registrar_app(self, request):
        user = request.user

        # Permitir: Director, Subdirector, Presidente y Jefe de Sección
        if user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE', 'JEFE_SECCION'] and not user.is_superuser:
            return Response({'error': 'Sin permisos. Solo Directores, Subdirectores, Presidentes y Jefes de Sección pueden registrar descuentos.'}, status=status.HTTP_403_FORBIDDEN)

        # Si es Jefe de Sección, obtener su sección
        seccion_permitida = None
        if user.rol == 'JEFE_SECCION':
            try:
                jefe = JefeSeccion.objects.get(musico__usuario=user)
                seccion_permitida = jefe.seccion
            except JefeSeccion.DoesNotExist:
                return Response({'error': 'No estás asignado como Jefe de Sección'}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data
        musicos = payload.get('musicos', [])
        if not isinstance(musicos, list) or not musicos:
            return Response({'error': 'La lista de músicos es obligatoria'}, status=status.HTTP_400_BAD_REQUEST)

        seccion = payload.get('seccion', '')
        referencia = payload.get('referencia', '')
        observaciones = payload.get('observaciones', 'Registro de descuentos desde app móvil')
        origen = 'APP_MOVIL'

        descuentos_objs = []
        errores = []

        for item in musicos:
            musico = None
            if item.get('musico_id'):
                musico = Musico.objects.filter(id=item['musico_id']).first()
            elif item.get('documento_identidad'):
                musico = Musico.objects.filter(documento_identidad=item['documento_identidad']).first()
            elif item.get('nombre'):
                partes = item['nombre'].strip().split()
                qs = Musico.objects.all()
                if len(partes) >= 2:
                    qs = qs.filter(nombres__icontains=partes[0], apellidos__icontains=partes[-1])
                else:
                    qs = qs.filter(nombres__icontains=partes[0])
                musico = qs.first()

            if not musico:
                errores.append({'item': item, 'error': 'Músico no encontrado'})
                continue

            # Si es Jefe de Sección, verificar que el músico sea de su sección
            if seccion_permitida and musico.instrumento != seccion_permitida:
                errores.append({'item': item, 'error': f'Músico no pertenece a tu sección ({seccion_permitida})'})
                continue

            monto = Decimal(str(item.get('monto', 0) or 0))
            if monto <= 0:
                errores.append({'item': item, 'error': 'Monto inválido'})
                continue

            fecha_valor = item.get('fecha')
            if fecha_valor:
                try:
                    fecha_falta = date.fromisoformat(fecha_valor)
                except Exception:
                    fecha_falta = date.today()
            else:
                fecha_falta = date.today()

            # Obtener jefe de sección si existe
            jefe_seccion = None
            if user.rol == 'JEFE_SECCION':
                try:
                    jefe_seccion = JefeSeccion.objects.get(musico__usuario=user)
                except JefeSeccion.DoesNotExist:
                    pass

            descuento = Descuento.objects.create(
                musico=musico,
                jefe_seccion=jefe_seccion,
                monto=monto,
                motivo=item.get('motivo', observaciones),
                fecha_falta=fecha_falta,
                origen=origen,
                estado='APROBADA'
            )
            descuentos_objs.append(descuento)

        response_data = {
            'success': True,
            'registrados': len(descuentos_objs),
            'errores': errores,
            'total_descuentos': sum(d.monto for d in descuentos_objs),
        }

        return Response(response_data)

class PagoViewSet(viewsets.ModelViewSet):
    queryset = Pago.objects.all().order_by('-fecha_liquidacion')
    serializer_class = PagoSerializer
    permission_classes = [permissions.IsAuthenticated]

class RankingView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        limite = int(request.query_params.get('limite', 10))
        calculator = RendimientoCalculator()
        ranking = calculator.generar_ranking_completo()
        
        return Response({
            'top_musicos': ranking['top_musicos'][:limite],
            'estadisticas': ranking['estadisticas'],
            'fecha_corte': ranking['fecha_corte']
        })

class ConfiguracionViewSet(viewsets.ModelViewSet):
    queryset = ConfiguracionSistema.objects.all()
    serializer_class = ConfiguracionSistemaSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_object(self):
        obj, created = ConfiguracionSistema.objects.get_or_create(id=1)
        return obj

class AdelantoViewSet(viewsets.ModelViewSet):
    queryset = Adelanto.objects.all().order_by('-fecha')
    serializer_class = AdelantoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrar adelantos según rol del usuario"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # DIRECTOR, SUBDIRECTOR, PRESIDENTE: ven todo
        if user.rol in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] or user.is_superuser:
            return queryset
        
        # JEFE DE SECCIÓN: solo adelantos de su sección
        if user.rol == 'JEFE_SECCION':
            try:
                jefe = JefeSeccion.objects.get(usuario=user)
                # Filtrar por sección del jefe
                return queryset.filter(musico__instrumento=jefe.seccion)
            except JefeSeccion.DoesNotExist:
                return queryset.none()
        
        # Otros roles: no ven adelantos
        return queryset.none()


    @action(detail=False, methods=['post'])
    def registrar_app(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data
        musicos = payload.get('musicos', [])
        if not isinstance(musicos, list) or not musicos:
            return Response({'error': 'La lista de músicos es obligatoria'}, status=status.HTTP_400_BAD_REQUEST)

        origen = 'APP'
        referencia = payload.get('referencia', '')
        observaciones = payload.get('observaciones', 'Registro desde app móvil')

        adelantos_objs = []
        errores = []
        for item in musicos:
            musico = None
            if item.get('musico_id'):
                musico = Musico.objects.filter(id=item['musico_id']).first()
            elif item.get('documento_identidad'):
                musico = Musico.objects.filter(documento_identidad=item['documento_identidad']).first()
            elif item.get('nombre'):
                partes = item['nombre'].strip().split()
                qs = Musico.objects.all()
                if len(partes) >= 2:
                    qs = qs.filter(nombres__icontains=partes[0], apellidos__icontains=partes[-1])
                else:
                    qs = qs.filter(nombres__icontains=partes[0])
                musico = qs.first()

            if not musico:
                errores.append({'item': item, 'error': 'Músico no encontrado'})
                continue

            monto = Decimal(str(item.get('monto', 0) or 0))
            if monto <= 0:
                errores.append({'item': item, 'error': 'Monto inválido'})
                continue

            fecha_valor = item.get('fecha')
            if fecha_valor:
                try:
                    fecha = date.fromisoformat(fecha_valor)
                except Exception:
                    fecha = date.today()
            else:
                fecha = date.today()

            adelanto = Adelanto.objects.create(
                musico=musico,
                monto=monto,
                motivo=item.get('motivo', observaciones),
                fecha=fecha,
                origen=origen
            )
            adelantos_objs.append(adelanto)

        response_data = {
            'success': True,
            'registrados': len(adelantos_objs),
            'errores': errores,
            'total_adelantos': sum(a.monto for a in adelantos_objs),
        }

        return Response(response_data)

class PlanillaLiquidacionViewSet(viewsets.ModelViewSet):
    queryset = PlanillaLiquidacion.objects.all().order_by('-fecha_creacion')
    serializer_class = PlanillaLiquidacionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PlanillaLiquidacionDetalleSerializer
        return super().get_serializer_class()

    @action(detail=False, methods=['post'])
    def update_order(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        order_data = request.data.get('order', [])
        if not order_data:
            return Response({'error': 'Datos de orden requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            for item in order_data:
                Musico.objects.filter(id=item['id']).update(orden=item['order'])
            return Response({'success': True})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def previsualizar(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)

        """
        Pre-visualiza la planilla de liquidación para una lista de eventos específicos.
        Recibe una lista de IDs de eventos y devuelve los cálculos para cada músico.
        """
        eventos_ids = request.data.get('eventos_ids', [])
        
        if not eventos_ids:
            return Response({'error': 'Debe proporcionar al menos un ID de evento'}, status=status.HTTP_400_BAD_REQUEST)
            
        eventos = Evento.objects.filter(id__in=eventos_ids)
        if eventos.count() != len(eventos_ids):
            return Response({'error': 'Uno o más eventos no existen'}, status=status.HTTP_400_BAD_REQUEST)

        # Obtener todas las asistencias no liquidadas para estos eventos
        asistencias = Asistencia.objects.filter(
            evento_id__in=eventos_ids,
            liquidado=False,
            estado__in=['PRESENTE', 'TARDANZA', 'JUSTIFICADO']
        ).select_related('musico', 'evento')

        # Agrupar por músico
        musicos_data = {}
        for asistencia in asistencias:
            musico_id = asistencia.musico.id
            if musico_id not in musicos_data:
                musicos_data[musico_id] = {
                    'musico': asistencia.musico,
                    'asistencias': [],
                    'monto_base': Decimal('0.00')
                }
            musicos_data[musico_id]['asistencias'].append(asistencia)
            musicos_data[musico_id]['monto_base'] += asistencia.monto_acordado

        resultados = []
        for musico_id, data in musicos_data.items():
            musico = data['musico']
            
            # Obtener descuentos y adelantos no liquidados del músico
            descuentos = Descuento.objects.filter(
                musico=musico, 
                liquidado=False
            )
            adelantos = Adelanto.objects.filter(
                musico=musico, 
                liquidado=False
            )

            total_descuentos = descuentos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            total_adelantos = adelantos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            
            monto_final = data['monto_base'] - total_descuentos - total_adelantos

            resultados.append({
                'musico_id': musico.id,
                'musico_nombre': musico.nombre_completo,
                'monto_base': data['monto_base'],
                'total_descuentos': total_descuentos,
                'total_adelantos': total_adelantos,
                'monto_final': monto_final,
                'detalle_asistencias': [
                    {
                        'evento_id': a.evento.id,
                        'evento_titulo': a.evento.titulo,
                        'monto': a.monto_acordado,
                        'estado': a.estado
                    } for a in data['asistencias']
                ],
                'detalle_descuentos': [
                    {
                        'id': d.id,
                        'motivo': d.motivo,
                        'monto': d.monto,
                        'fecha_falta': d.fecha_falta
                    } for d in descuentos
                ],
                'detalle_adelantos': [
                    {
                        'id': ad.id,
                        'motivo': ad.motivo,
                        'monto': ad.monto,
                        'fecha': ad.fecha
                    } for ad in adelantos
                ],
            })

        # Ordenar por nombre de músico
        resultados.sort(key=lambda x: x['musico_nombre'])
        
        return Response({
            'eventos': [{'id': e.id, 'titulo': e.titulo} for e in eventos],
            'musicos': resultados,
            'resumen': {
                'total_musicos': len(resultados),
                'total_monto_base': sum(r['monto_base'] for r in resultados),
                'total_descuentos': sum(r['total_descuentos'] for r in resultados),
                'total_adelantos': sum(r['total_adelantos'] for r in resultados),
                'total_a_pagar': sum(r['monto_final'] for r in resultados)
            }
        })

    @action(detail=False, methods=['get'])
    def simular(self, request):
        musico_id = request.query_params.get('musico_id')
        
        musicos = Musico.objects.filter(activo=True)
        if musico_id:
            musicos = musicos.filter(id=musico_id)

        resultados = []
        for musico in musicos:
            asistencias = Asistencia.objects.filter(musico=musico, liquidado=False, estado__in=['PRESENTE', 'TARDANZA', 'JUSTIFICADO'])
            descuentos = Descuento.objects.filter(musico=musico, estado='APROBADA')
            adelantos = Adelanto.objects.filter(musico=musico, estado='APROBADA')

            monto_base = asistencias.aggregate(total=Sum('monto_acordado'))['total'] or Decimal('0.00')
            total_descuentos = descuentos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            total_adelantos = adelantos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')

            monto_final = monto_base - total_descuentos - total_adelantos

            if monto_base > 0 or total_descuentos > 0 or total_adelantos > 0:
                resultados.append({
                    'musico_id': musico.id,
                    'musico_nombre': musico.nombre_completo,
                    'monto_base': monto_base,
                    'total_descuentos': total_descuentos,
                    'total_adelantos': total_adelantos,
                    'monto_final': monto_final,
                    'detalle_asistencias': [{'evento': a.evento.titulo, 'monto': a.monto_acordado} for a in asistencias],
                    'detalle_descuentos': [{'motivo': d.motivo, 'monto': d.monto} for d in descuentos],
                    'detalle_adelantos': [{'motivo': ad.motivo, 'monto': ad.monto} for ad in adelantos],
                })

        return Response(resultados)

    @action(detail=False, methods=['post'])
    def liquidar_directo(self, request):
        """
        Liquida un evento directamente desde una tabla editable donde el usuario introduce
        los montos totales acordados, multas y adelantos manualmente o desde PDF.
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        evento_id = request.data.get('evento_id')
        titulo = request.data.get('titulo', f'Planilla {date.today().strftime("%Y-%m-%d")}')
        observaciones = request.data.get('observaciones', '')
        datos_musicos = request.data.get('musicos', []) # Lista de { musico_id, acordado, multas, adelantos }
        
        if not evento_id:
            return Response({'error': 'Debe seleccionar un evento'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            evento = Evento.objects.get(id=evento_id)
        except Evento.DoesNotExist:
            return Response({'error': 'El evento no existe'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear la planilla de liquidación
        planilla = PlanillaLiquidacion.objects.create(
            titulo=titulo,
            observaciones=observaciones,
            completada=True
        )
        planilla.eventos.add(evento)

        pagos_generados = 0
        for dato in datos_musicos:
            try:
                musico = Musico.objects.get(id=dato['musico_id'])
            except Musico.DoesNotExist:
                continue

            acordado = Decimal(str(dato.get('acordado', 0) or 0))
            multas = Decimal(str(dato.get('multas', 0) or 0))
            adelantos = Decimal(str(dato.get('adelantos', 0) or 0))
            descuentos_extra = dato.get('descuentos_extra', [])

            total_extra = Decimal('0.00')
            for extra in descuentos_extra:
                monto_extra = Decimal(str(extra.get('monto', 0) or 0))
                if monto_extra > 0:
                    total_extra += monto_extra
                    Descuento.objects.create(
                        musico=musico,
                        monto=monto_extra,
                        motivo=f"{extra.get('nombre')} en {evento.titulo}",
                        fecha_falta=date.today(),
                        liquidado=True
                    )

            monto_final = acordado - multas - adelantos - total_extra

            # Registrar Asistencia/Acordado
            Asistencia.objects.update_or_create(
                musico=musico,
                evento=evento,
                defaults={
                    'monto_acordado': acordado,
                    'liquidado': True,
                    'estado': 'PRESENTE' # Asumimos presente si se le está pagando
                }
            )

            # Registrar histórico de Multa si hay
            if multas > 0:
                Descuento.objects.create(
                    musico=musico,
                    monto=multas,
                    motivo=f'Multa en {evento.titulo}',
                    fecha_falta=date.today(),
                    liquidado=True
                )

            # Registrar histórico de Adelanto si hay
            if adelantos > 0:
                Adelanto.objects.create(
                    musico=musico,
                    monto=adelantos,
                    motivo=f'Adelanto deducido en {evento.titulo}',
                    fecha=date.today(),
                    liquidado=True
                )

            # Crear el Pago final
            if acordado > 0 or multas > 0 or adelantos > 0:
                Pago.objects.create(
                    musico=musico,
                    planilla=planilla,
                    salario_base=acordado,
                    descuentos_totales=multas,
                    adelantos_totales=adelantos,
                    neto_pagar=monto_final,
                    fecha_liquidacion=timezone.now(),
                    estado='PENDIENTE',
                    observaciones=f"Liquidación directa - {evento.titulo}"
                )
                pagos_generados += 1

        return Response({
            'success': True, 
            'planilla_id': planilla.id, 
            'planilla_titulo': planilla.titulo,
            'pagos_generados': pagos_generados
        })

    @action(detail=False, methods=['post'])
    def liquidar_app(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)

        evento_id = request.data.get('evento_id')
        titulo = request.data.get('titulo', f'Liquidación App {date.today().strftime("%Y-%m-%d")}')
        observaciones = request.data.get('observaciones', '')
        musicos = request.data.get('musicos', [])

        if not evento_id:
            return Response({'error': 'Debe seleccionar un evento'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(musicos, list) or not musicos:
            return Response({'error': 'La lista de músicos es obligatoria'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            evento = Evento.objects.get(id=evento_id)
        except Evento.DoesNotExist:
            return Response({'error': 'El evento no existe'}, status=status.HTTP_400_BAD_REQUEST)

        planilla = PlanillaLiquidacion.objects.create(
            titulo=titulo,
            observaciones=observaciones,
            completada=True
        )
        planilla.eventos.add(evento)

        pagos_generados = []
        for item in musicos:
            musico = None
            if item.get('musico_id'):
                musico = Musico.objects.filter(id=item['musico_id']).first()
            elif item.get('documento_identidad'):
                musico = Musico.objects.filter(documento_identidad=item['documento_identidad']).first()
            if not musico:
                continue

            monto_base = Decimal(str(item.get('monto_base', 0) or 0))
            descuentos = item.get('descuentos', [])
            adelantos = item.get('adelantos', [])

            total_descuentos = Decimal('0.00')
            for d in descuentos:
                monto_desc = Decimal(str(d.get('monto', 0) or 0))
                if monto_desc > 0:
                    total_descuentos += monto_desc
                    fecha_descuento = date.today()
                    if d.get('fecha_falta'):
                        try:
                            fecha_descuento = date.fromisoformat(d.get('fecha_falta'))
                        except Exception:
                            pass
                    Descuento.objects.create(
                        musico=musico,
                        monto=monto_desc,
                        motivo=d.get('motivo', f'Descuento app para {evento.titulo}'),
                        fecha_falta=fecha_descuento,
                        origen='APP'
                    )

            total_adelantos = Decimal('0.00')
            for a in adelantos:
                monto_adel = Decimal(str(a.get('monto', 0) or 0))
                if monto_adel > 0:
                    total_adelantos += monto_adel
                    fecha_adelanto = date.today()
                    if a.get('fecha'):
                        try:
                            fecha_adelanto = date.fromisoformat(a.get('fecha'))
                        except Exception:
                            pass
                    Adelanto.objects.create(
                        musico=musico,
                        monto=monto_adel,
                        motivo=a.get('motivo', f'Adelanto app para {evento.titulo}'),
                        fecha=fecha_adelanto,
                        origen='APP'
                    )

            monto_final = monto_base - total_descuentos - total_adelantos

            pago = Pago.objects.create(
                musico=musico,
                planilla=planilla,
                salario_base=monto_base,
                descuentos_totales=total_descuentos,
                adelantos_totales=total_adelantos,
                neto_pagar=monto_final,
                fecha_liquidacion=timezone.now(),
                estado='PENDIENTE',
                observaciones=f'Liquidación desde app - {observaciones}'
            )
            pagos_generados.append({
                'musico_id': musico.id,
                'musico_nombre': musico.nombre_completo,
                'pago_id': pago.id,
                'monto_final': monto_final,
            })

        return Response({
            'success': True,
            'planilla_id': planilla.id,
            'planilla_titulo': planilla.titulo,
            'pagos': pagos_generados,
            'total_pagos': sum(p['monto_final'] for p in pagos_generados)
        })

        """
        Consolida/paga la planilla de liquidación para una lista de eventos específicos.
        Crea la planilla, los pagos y marca todos los registros como liquidados.
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        eventos_ids = request.data.get('eventos_ids', [])
        titulo = request.data.get('titulo', f'Planilla {date.today().strftime("%Y-%m-%d")}')
        observaciones = request.data.get('observaciones', '')
        
        if not eventos_ids:
            return Response({'error': 'Debe proporcionar al menos un ID de evento'}, status=status.HTTP_400_BAD_REQUEST)
            
        eventos = Evento.objects.filter(id__in=eventos_ids)
        if eventos.count() != len(eventos_ids):
            return Response({'error': 'Uno o más eventos no existen'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear la planilla de liquidación
        planilla = PlanillaLiquidacion.objects.create(
            titulo=titulo,
            observaciones=observaciones,
            completada=True
        )
        
        # Agregar los eventos a la planilla
        planilla.eventos.add(*eventos_ids)

        # Obtener todas las asistencias no liquidadas para estos eventos
        asistencias = Asistencia.objects.filter(
            evento_id__in=eventos_ids,
            liquidado=False,
            estado__in=['PRESENTE', 'TARDANZA', 'JUSTIFICADO']
        ).select_related('musico', 'evento')

        # Agrupar por músico
        musicos_data = {}
        for asistencia in asistencias:
            musico_id = asistencia.musico.id
            if musico_id not in musicos_data:
                musicos_data[musico_id] = {
                    'musico': asistencia.musico,
                    'asistencias': [],
                    'monto_base': Decimal('0.00')
                }
            musicos_data[musico_id]['asistencias'].append(asistencia)
            musicos_data[musico_id]['monto_base'] += asistencia.monto_acordado

        pagos_generados = 0
        for musico_id, data in musicos_data.items():
            musico = data['musico']
            
            # Obtener descuentos y adelantos no liquidados del músico
            descuentos = Descuento.objects.filter(
                musico=musico, 
                liquidado=False
            )
            adelantos = Adelanto.objects.filter(
                musico=musico, 
                liquidado=False
            )

            total_descuentos = descuentos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            total_adelantos = adelantos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            
            monto_final = data['monto_base'] - total_descuentos - total_adelantos

            # Solo crear pago si hay algo que pagar (positivo o negativo)
            if data['monto_base'] > 0 or total_descuentos > 0 or total_adelantos > 0:
                pago = Pago.objects.create(
                    musico=musico,
                    planilla=planilla,
                    salario_base=data['monto_base'],
                    descuentos_totales=total_descuentos,
                    adelantos_totales=total_adelantos,
                    neto_pagar=monto_final,
                    fecha_liquidacion=timezone.now(),
                    estado='PENDIENTE',
                    observaciones=f"Liquidación automática - {titulo}"
                )
                
                # Marcar descuentos y adelantos como liquidados
                descuentos.update(estado='LIQUIDADA')
                adelantos.update(estado='LIQUIDADA')
                
                pagos_generados += 1

        return Response({
            'success': True, 
            'planilla_id': planilla.id, 
            'planilla_titulo': planilla.titulo,
            'pagos_generados': pagos_generados,
            'eventos_procesados': len(eventos_ids),
            'fecha_creacion': planilla.fecha_creacion
        })

    @action(detail=True, methods=['get'])
    def generar_pdf(self, request, pk=None):
        """
        Genera un PDF detallado de la planilla de liquidación.
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        planilla = self.get_object()
        
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from reportlab.lib.units import inch
        except ImportError:
            return Response({'error': 'reportlab no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        import io
        from django.http import HttpResponse
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=70)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
            fontSize=16, textColor=colors.black, alignment=TA_CENTER,
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'], fontName='Helvetica-Bold',
            fontSize=12, textColor=colors.black, alignment=TA_CENTER,
            spaceAfter=20
        )
        normal_style = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontSize=9)
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10)
        
        # Encabezado
        elements.append(Paragraph("BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA<br/>Eco De Los Andes", title_style))
        elements.append(Paragraph(f"PLANILLA DE LIQUIDACIÓN - {planilla.titulo.upper()}", subtitle_style))
        
        if planilla.observaciones:
            elements.append(Paragraph(f"<b>Observaciones:</b> {planilla.observaciones}", normal_style))
        
        elements.append(Paragraph(f"<b>Fecha de Creación:</b> {planilla.fecha_creacion.strftime('%d/%m/%Y %H:%M')}", normal_style))
        elements.append(Spacer(1, 20))
        
        # Tabla de eventos
        if planilla.eventos.exists():
            elements.append(Paragraph("<b>EVENTOS INCLUIDOS:</b>", header_style))
            eventos_data = [['N°', 'Evento', 'Fecha']]
            for i, evento in enumerate(planilla.eventos.all(), 1):
                eventos_data.append([
                    str(i),
                    evento.titulo,
                    evento.fecha_hora_cita.strftime('%d/%m/%Y')
                ])
            
            eventos_table = Table(eventos_data)
            eventos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(eventos_table)
            elements.append(Spacer(1, 20))
        
        # Tabla de pagos
        if planilla.pagos.exists():
            elements.append(Paragraph("<b>DETALLE DE PAGOS:</b>", header_style))
            pagos_data = [
                ['N°', 'Músico', 'Monto Base', 'Descuentos', 'Adelantos', 'Monto Final', 'Estado']
            ]
            
            for i, pago in enumerate(planilla.pagos.all(), 1):
                pagos_data.append([
                    str(i),
                    pago.musico.nombre_completo,
                    f"${pago.monto_base:.2f}",
                    f"${pago.descuentos_aplicados:.2f}",
                    f"${pago.adelantos_aplicados:.2f}",
                    f"${pago.monto_final:.2f}",
                    'PAGADO' if pago.pagado else 'PENDIENTE'
                ])
            
            # Fila de totales
            total_base = sum(p.monto_base for p in planilla.pagos.all())
            total_descuentos = sum(p.descuentos_aplicados for p in planilla.pagos.all())
            total_adelantos = sum(p.adelantos_aplicados for p in planilla.pagos.all())
            total_final = sum(p.monto_final for p in planilla.pagos.all())
            
            pagos_data.append([
                '',
                '<b>TOTALES:</b>',
                f"<b>${total_base:.2f}</b>",
                f"<b>${total_descuentos:.2f}</b>",
                f"<b>${total_adelantos:.2f}</b>",
                f"<b>${total_final:.2f}</b>",
                ''
            ])
            
            pagos_table = Table(pagos_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
            pagos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFFF99')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(pagos_table)
        
        # Función para pie de página
        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(doc.pagesize[0] - 30, 20, f"Página {doc.page}")
            canvas.drawString(30, 20, f"Generado por: {request.user.get_full_name() or request.user.username}")
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Planilla_{planilla.titulo.replace(" ", "_")}.pdf"'
        return response

    @action(detail=False, methods=['get'])
    def reporte_general(self, request):
        """
        Genera un PDF con el reporte general de finanzas.
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            return Response({'error': 'reportlab no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        import io
        from django.http import HttpResponse
        from datetime import datetime
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=70)
        elements = []
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
            fontSize=16, textColor=colors.black, alignment=TA_CENTER,
            spaceAfter=10
        )
        normal_style = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontSize=9)
        
        # Encabezado
        elements.append(Paragraph("BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA<br/>Eco De Los Andes", title_style))
        elements.append(Paragraph(f"REPORTE GENERAL DE FINANZAS<br/>{datetime.now().strftime('%d/%m/%Y')}", title_style))
        elements.append(Spacer(1, 20))
        
        # Resumen general
        resumen_data = [
            ['Métrica', 'Total'],
            ['Músicos Activos', str(Musico.objects.filter(activo=True).count())],
            ['Total Eventos', str(Evento.objects.count())],
            ['Total Asistencias', str(Asistencia.objects.count())],
            ['Total Pagos', str(Pago.objects.count())],
            ['Planillas Creadas', str(PlanillaLiquidacion.objects.count())],
        ]
        
        resumen_table = Table(resumen_data)
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(resumen_table)
        elements.append(Spacer(1, 20))
        
        # Pagos pendientes (solo referencia informativa)
        pagos_pendientes = Pago.objects.filter(pagado=False)
        if pagos_pendientes.exists():
            elements.append(Paragraph("<b>PAGOS PENDIENTES:</b>", normal_style))
            pendientes_data = [['Músico', 'Monto', 'Fecha']]
            for pago in pagos_pendientes:
                pendientes_data.append([
                    pago.musico.nombre_completo,
                    f"${pago.monto_final:.2f}",
                    pago.fecha_pago.strftime('%d/%m/%Y')
                ])
            
            pendientes_table = Table(pendientes_data)
            pendientes_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFB6C1')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(pendientes_table)
        
        # Pie de página
        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(doc.pagesize[0] - 30, 20, f"Página {doc.page}")
            canvas.drawString(30, 20, f"Generado por: {request.user.get_full_name() or request.user.username}")
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_General_Finanzas.pdf"'
        return response

class ContratoMusicoViewSet(viewsets.ModelViewSet):
    queryset = ContratoMusico.objects.all().order_by('-created_at')
    serializer_class = None
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        from .serializers import ContratoMusicoSerializer
        return ContratoMusicoSerializer
    
    @action(detail=False, methods=['post'])
    def asignar_montos_personalizados(self, request):
        """
        Permite al directorio asignar montos personalizados para un evento
        Soporta montos variables por día
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
        
        evento_id = request.data.get('evento_id')
        contratos_data = request.data.get('contratos', [])
        
        if not evento_id or not contratos_data:
            return Response({'error': 'Datos incompletos'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            evento = Evento.objects.get(id=evento_id)
        except Evento.DoesNotExist:
            return Response({'error': 'Evento no existe'}, status=status.HTTP_404_NOT_FOUND)
        
        creados = 0
        actualizados = 0
        detalles_creados = 0
        
        for contrato_data in contratos_data:
            musico_id = contrato_data.get('musico_id')
            monto_diario_base = contrato_data.get('monto_diario')
            montos_diarios = contrato_data.get('montos_diarios', [])  # Lista de montos por fecha
            observaciones = contrato_data.get('observaciones', '')
            
            try:
                musico = Musico.objects.get(id=musico_id)
            except Musico.DoesNotExist:
                continue
            
            # Crear o actualizar contrato base
            contrato, created = ContratoMusico.objects.update_or_create(
                musico=musico,
                evento=evento,
                defaults={
                    'monto_diario': monto_diario_base,
                    'aprobado_por': request.user,
                    'fecha_aprobacion': timezone.now(),
                    'observaciones': observaciones
                }
            )
            
            if created:
                creados += 1
            else:
                actualizados += 1
            
            # Procesar montos diarios específicos si existen
            if montos_diarios:
                for monto_diario_data in montos_diarios:
                    fecha = monto_diario_data.get('fecha')
                    monto_especificico = monto_diario_data.get('monto')
                    motivo = monto_diario_data.get('motivo', '')
                    
                    if fecha and monto_especificico:
                        from datetime import datetime
                        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                        
                        detalle, detalle_created = DetalleMontoDiario.objects.update_or_create(
                            contrato=contrato,
                            fecha=fecha_obj,
                            defaults={
                                'monto_asignado': monto_especificico,
                                'motivo_variacion': motivo,
                                'aprobado_por': request.user,
                                'fecha_aprobacion': timezone.now()
                            }
                        )
                        
                        if detalle_created:
                            detalles_creados += 1
                        
                        # Crear o actualizar asistencia para esa fecha específica
                        Asistencia.objects.update_or_create(
                            musico=musico,
                            evento=evento,
                            fecha_asistencia=fecha_obj,
                            defaults={
                                'monto_acordado': monto_especificico,
                                'contrato': contrato
                            }
                        )
            else:
                # Si no hay montos diarios específicos, usar el monto base para la fecha del evento
                fecha_evento = evento.fecha_hora_cita.date()
                Asistencia.objects.update_or_create(
                    musico=musico,
                    evento=evento,
                    fecha_asistencia=fecha_evento,
                    defaults={
                        'monto_acordado': monto_diario_base,
                        'contrato': contrato
                    }
                )
        
        return Response({
            'success': True,
            'creados': creados,
            'actualizados': actualizados,
            'detalles_diarios': detalles_creados,
            'evento': evento.titulo
        })

    @action(detail=False, methods=['post'])
    def asignar_montos_variables(self, request):
        """
        Asigna montos variables para músicos en diferentes fechas de un mismo evento
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
        
        evento_id = request.data.get('evento_id')
        musico_id = request.data.get('musico_id')
        montos_por_fecha = request.data.get('montos_por_fecha', [])
        
        if not evento_id or not musico_id or not montos_por_fecha:
            return Response({'error': 'Datos incompletos'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            evento = Evento.objects.get(id=evento_id)
            musico = Musico.objects.get(id=musico_id)
        except (Evento.DoesNotExist, Musico.DoesNotExist):
            return Response({'error': 'Evento o músico no existe'}, status=status.HTTP_404_NOT_FOUND)
        
        # Obtener o crear contrato base
        contrato, created = ContratoMusico.objects.get_or_create(
            musico=musico,
            evento=evento,
            defaults={
                'monto_diario': Decimal('0.00'),
                'aprobado_por': request.user,
                'fecha_aprobacion': timezone.now()
            }
        )
        
        detalles_procesados = 0
        for monto_data in montos_por_fecha:
            fecha = monto_data.get('fecha')
            monto = monto_data.get('monto')
            motivo = monto_data.get('motivo', '')
            
            if fecha and monto:
                from datetime import datetime
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                
                # Crear detalle diario
                detalle, detalle_created = DetalleMontoDiario.objects.update_or_create(
                    contrato=contrato,
                    fecha=fecha_obj,
                    defaults={
                        'monto_asignado': monto,
                        'motivo_variacion': motivo,
                        'aprobado_por': request.user,
                        'fecha_aprobacion': timezone.now()
                    }
                )
                
                if detalle_created:
                    detalles_procesados += 1
                
                # Crear asistencia específica
                Asistencia.objects.update_or_create(
                    musico=musico,
                    evento=evento,
                    fecha_asistencia=fecha_obj,
                    defaults={
                        'monto_acordado': monto,
                        'contrato': contrato
                    }
                )
        
        return Response({
            'success': True,
            'detalles_procesados': detalles_procesados,
            'musico': musico.nombre_completo,
            'evento': evento.titulo
        })

class DeudaViewSet(viewsets.ModelViewSet):
    queryset = Deuda.objects.all().order_by('-fecha_creacion')
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Deuda.objects.all().order_by('-fecha_creacion')
        musico_id = self.request.query_params.get('musico_id')
        seccion = self.request.query_params.get('seccion')

        if musico_id:
            queryset = queryset.filter(musico_id=musico_id)
        if seccion:
            queryset = queryset.filter(musico__instrumento__iexact=seccion)

        return queryset

    def get_serializer_class(self):
        from .serializers import DeudaSerializer
        return DeudaSerializer

    @action(detail=False, methods=['post'])
    def crear_masivo(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        motivo = request.data.get('motivo')
        monto_total = request.data.get('monto_total')
        musicos_ids = request.data.get('musicos_ids', [])
        
        if not motivo or not monto_total or not musicos_ids:
            return Response({'error': 'Faltan datos requeridos'}, status=status.HTTP_400_BAD_REQUEST)
            
        creadas = 0
        for m_id in musicos_ids:
            try:
                musico = Musico.objects.get(id=m_id)
                Deuda.objects.create(
                    musico=musico,
                    motivo=motivo,
                    monto_total=monto_total
                )
                creadas += 1
            except Musico.DoesNotExist:
                pass
                
        return Response({'success': True, 'creadas': creadas})

class AbonoDeudaViewSet(viewsets.ModelViewSet):
    queryset = AbonoDeuda.objects.all().order_by('-fecha')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        from .serializers import AbonoDeudaSerializer
        return AbonoDeudaSerializer

class JefeSeccionViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar jefes de sección"""
    queryset = JefeSeccion.objects.all().order_by('-fecha_nombramiento')
    serializer_class = None  # Se definirá dinámicamente
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['seccion', 'activo']
    search_fields = ['musico__nombre_completo', 'seccion']
    ordering_fields = ['fecha_nombramiento', 'seccion', 'musico__nombre_completo']

    def get_serializer_class(self):
        from .serializers import JefeSeccionSerializer
        return JefeSeccionSerializer

    @action(detail=False, methods=['get'])
    def por_seccion(self, request):
        """Retorna los jefes activos agrupados por sección"""
        jefes = JefeSeccion.objects.filter(activo=True).select_related('musico')
        serializer = self.get_serializer(jefes, many=True)
        return Response(serializer.data)

